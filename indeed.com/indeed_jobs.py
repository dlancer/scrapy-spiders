# -*- coding: utf-8 -*-
# (c) dlancer, 2017

from scrapy import Spider, Request

from openpyxl import load_workbook


BASE_URL = 'https://www.indeed.com'
LOCATIONS_FILENAME = ''
SHEET_NAME = 'locations'


class IndeedJobsSpider(Spider):
    name = 'indeed_jobs'
    allowed_domains = ['indeed.com']

    def __init__(self, query=None, city=None, state=None, job_type=None, radius=None, *args, **kwargs):
        super(IndeedJobsSpider, self).__init__(*args, **kwargs)
        self.query = query
        self.city = city if city is not None else 'New+York'
        self.state = state if state is not None else 'NY'
        self.job_type = job_type if job_type is not None else 'fulltime'
        self.radius = radius

    def start_requests(self):

        #  location reading from xlsx file
        try:
            wb = load_workbook(filename=LOCATIONS_FILENAME)
            sheet_ranges = wb[SHEET_NAME]
            locations = []
            for row in sheet_ranges.iter_rows(row_offset=1):
                if row[0].value is not None:
                    locations.append((row[0].value, str(row[1].value)))
            self.logger.info('Locations loaded from file')
        except IOError:
            locations = None

        if self.query is None:
            self.logger.error('You should add query for job search')
            return

        if locations is None:
            url = BASE_URL + '/jobs?q={0}&l={1},+{2}&jt={3}'.format(
                self.query,
                self.city,
                self.state,
                self.job_type,
            )
            if self.radius is not None:
                url += '&radius={0}'.format(self.radius)
            urls = [url]
        else:
            urls = []
            # zip is not used now
            for location in locations:
                self.city, self.state = location[0].encode('utf-8').split(',')
                self.city = self.city.lstrip('\n\t ').rstrip('\n\t ').replace(' ', '+')
                self.state = self.state.lstrip('\n\t ').rstrip('\n\t ')
                url = BASE_URL + '/jobs?q={0}&l={1},+{2}&jt={3}'.format(
                    self.query,
                    self.city,
                    self.state,
                    self.job_type,
                )
                if self.radius is not None:
                    url += '&radius={0}'.format(self.radius)
                urls.append(url)

        # URL examples:
        # 'https://www.indeed.com/jobs?q=Java+$100,000&l=New+York,+NY&jt=contract',
        # 'https://www.indeed.com/jobs?q=Hadoop+$100,000&l=New+York,+NY&jt=fulltime',
        # 'https://www.indeed.com/jobs?q=Hadoop+$100,000&jt=fulltime&l=New+York,+NY&radius=100',

        # TODO: fix url encodings?
        for url in urls:
            yield Request(url=url, callback=self.parse)

    def parse(self, response):

        # if proxy is enabled check proper response from site
        if self.settings.get('PROXY_USE'):
            if not response.xpath('//img[@class="indeedLogo"]'):
                yield Request(url=response.url, dont_filter=True, callback=self.parse)

        # item parsing
        for job in response.css('div.row').css('.result'):
            job_title = job.css('h2.jobtitle a::text').extract_first().lstrip('\n\t ')
            company = job.xpath('//span[@class="company"]/span/text()').extract_first()
            if company is None:
                url = job.xpath('//span[@class="company"]/span/a/@href').extract_first()
                job_company_url = BASE_URL + url
                job_company = job.xpath('//span[@class="company"]/span/a/text()').extract_first().lstrip('\n\t ')
            else:
                job_company_url = ''
                job_company = company.lstrip('\n\t ')
            location = job.css('span.location span::text').extract()
            city, state = location[0].lstrip('\n\t ').split(',')
            state = state.lstrip('\n\t ').rstrip('\n\t ')
            try:
                state, zipc = state.split(' ')
            except ValueError:
                pass

            if len(location) > 1:
                location_info = location[1].lstrip('\n\t ').replace('(', '').replace(')', '')
            else:
                location_info = ''
            job_description = job.css('span.summary::text').extract_first().lstrip('\n\t ')
            job_salary = job.xpath('//td[@class="snip"]/nobr/text()').extract_first()
            job_salary = '' if job_salary is None else job_salary.strip('\n\t')

            yield {
                'job_title': job_title,
                'company': job_company,
                'company_url': job_company_url,
                'city': city,
                'state': state,
                'location_info': location_info,
                'salary': job_salary,
                'job_description': job_description
            }

        # select the next page for parsing
        next_page = response.xpath('//div[@class="pagination"]//a[contains(.,"Next")][1]/@href').extract_first()
        if next_page is not None:
            next_url = BASE_URL + next_page
            yield Request(url=next_url, callback=self.parse)
